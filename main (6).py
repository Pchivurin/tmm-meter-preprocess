"""
Сервіс OCR лічильників для ТММ (A3: важка робота в сервісі, n8n тонкий).
Два ендпоінти:
  /locate — фото → {serial, apartment, x1,y1,x2,y2}  (легка відповідь, без фото)
  /read   — фото + координати + prev → {reading, confidence}  (вирізає дисплей, читає з якорем)

Claude викликається всередині сервісу (ключ у змінній ANTHROPIC_API_KEY).
Base64 назад не повертається — тільки текст. Це знімає OOM у n8n.
"""
import base64
import json
import os
import urllib.request

import cv2
import numpy as np
from fastapi import FastAPI, HTTPException
from pydantic import BaseModel

app = FastAPI(title="TMM Meter OCR", version="3.0")

ANTHROPIC_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
MODEL = "claude-sonnet-4-6"


def call_claude(system, image_b64, user_text, max_tokens=500):
    body = {
        "model": MODEL,
        "max_tokens": max_tokens,
        "system": system,
        "messages": [{"role": "user", "content": [
            {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": image_b64}},
            {"type": "text", "text": user_text},
        ]}],
    }
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=json.dumps(body).encode(),
        headers={
            "content-type": "application/json",
            "anthropic-version": "2023-06-01",
            "x-api-key": ANTHROPIC_KEY,
        },
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        data = json.loads(resp.read())
    txt = "".join(c.get("text", "") for c in data.get("content", []) if c.get("type") == "text")
    return txt


def extract_json(txt):
    txt = (txt or "").replace("```json", "").replace("```", "")
    a, b = txt.find("{"), txt.rfind("}")
    if a != -1 and b != -1 and b > a:
        try:
            return json.loads(txt[a:b + 1])
        except Exception:
            return {}
    return {}


def decode_img(b64):
    raw = base64.b64decode(b64)
    arr = np.frombuffer(raw, dtype=np.uint8)
    img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
    if img is None:
        raise ValueError("imdecode None")
    return img


# ---------- /locate ----------
class LocateRequest(BaseModel):
    image_base64: str


LOCATE_SYS = (
    "Ти аналізуєш фото електролічильника. Поверни: (1) серійний номер зі штрихкоду, "
    "(2) номер квартири якщо є, (3) КООРДИНАТИ ДИСПЛЕЯ (LCD-екран з великими цифрами показника, "
    "напр. 048457.99; сірий/зеленуватий прямокутник, часто вгорі ліворуч). "
    "НЕ паспортна табличка (дрібні 10(85)А, 50Гц, 3х220/380В), НЕ штрихкод, НЕ логотип. "
    "Координати у шкалі 0-1000: x1,y1 лівий-верхній, x2,y2 правий-нижній екрана. Обведи щільно цифри. "
    'Поверни СТРОГО JSON без markdown: {"serial":"<цифри або порожньо>","apartment":"<номер або порожньо>",'
    '"x1":<0-1000>,"y1":<0-1000>,"x2":<0-1000>,"y2":<0-1000>}'
)


@app.post("/locate")
def locate(req: LocateRequest):
    if not ANTHROPIC_KEY:
        raise HTTPException(500, "ANTHROPIC_API_KEY не заданий")
    try:
        txt = call_claude(LOCATE_SYS, req.image_base64,
                          "Знайди екран з цифрами показника (не паспортну табличку). Дай координати, серійний, квартиру.")
    except Exception as e:
        raise HTTPException(502, f"Claude locate error: {e}")
    p = extract_json(txt)
    def n(v, d):
        try:
            x = float(v)
            return x if 0 <= x <= 1000 else d
        except Exception:
            return d
    x1, y1, x2, y2 = n(p.get("x1"), 0), n(p.get("y1"), 0), n(p.get("x2"), 1000), n(p.get("y2"), 1000)
    if x2 <= x1 or y2 <= y1:
        x1, y1, x2, y2 = 0, 0, 1000, 1000
    return {"serial": p.get("serial", ""), "apartment": p.get("apartment", ""),
            "x1": x1, "y1": y1, "x2": x2, "y2": y2}


# ---------- /read ----------
class ReadRequest(BaseModel):
    image_base64: str
    x1: float; y1: float; x2: float; y2: float
    prev_reading: str = ""
    scale: float = 3.0
    clip_limit: float = 2.5


def crop_enhance(img, x1, y1, x2, y2, scale, clip):
    h, w = img.shape[:2]
    # Координати Locate нестабільні (стрибають). Замість тісного боксу
    # беремо ШИРОКУ зону навколо центру боксу з великим запасом,
    # гарантовано покриваючи дисплей навіть при промаху Locate.
    cx = (x1 + x2) / 2.0
    cy = (y1 + y2) / 2.0
    # напівширина/напіввисота зони: щедрі, щоб дисплей точно потрапив
    half_w = 380  # у шкалі 0-1000 -> покриває ~76% ширини
    half_h = 220  # покриває ~44% висоти навколо центру
    ex1 = max(0, cx - half_w)
    ex2 = min(1000, cx + half_w)
    ey1 = max(0, cy - half_h)
    ey2 = min(1000, cy + half_h)
    px1, py1 = int(ex1 / 1000 * w), int(ey1 / 1000 * h)
    px2, py2 = int(ex2 / 1000 * w), int(ey2 / 1000 * h)
    if px2 - px1 < 10 or py2 - py1 < 10:
        crop = img
    else:
        crop = img[py1:py2, px1:px2]
    crop = cv2.resize(crop, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
    lab = cv2.cvtColor(crop, cv2.COLOR_BGR2LAB)
    l, a, b = cv2.split(lab)
    l = cv2.createCLAHE(clipLimit=clip, tileGridSize=(8, 8)).apply(l)
    crop = cv2.cvtColor(cv2.merge([l, a, b]), cv2.COLOR_LAB2BGR)
    crop = cv2.filter2D(crop, -1, np.array([[0, -1, 0], [-1, 5, -1], [0, -1, 0]]))
    return crop


READ_SYS = (
    "Ти зчитуєш ПОКАЗНИК з крупного зображення табло електролічильника. Точність КРИТИЧНА.\n"
    "1. Головне число — показник активної енергії (kWh, A/A+).\n"
    "2. Читай цифра за цифрою зліва направо. Ціла частина зазвичай 5-6 цифр.\n"
    "3. ПРОМІЖКИ: якщо цифри розділені пробілом — це ОДНЕ число, з'єднай. Пробіл НЕ кома.\n"
    "4. {ANCHOR}"
    "5. ОСТАННЯ ЦИФРА цілої частини (перед комою) — НАЙЧАСТІША помилка. Прочитай її ОКРЕМО, двічі, дуже уважно.\n"
    "6. РОЗРІЗНЯЙ схожі семисегментні цифри за тим, які сегменти світяться:\n"
    "   - 0 vs 8: у 8 світяться ОБИДВА середні бічні сегменти (замкнена вісімка); у 0 середня риска НЕ світиться.\n"
    "   - 2 vs 9: у 9 світиться верхній правий контур із замкненою петлею вгорі; 2 має характерний злам.\n"
    "   - 1 vs 7: 7 має верхню горизонтальну риску; 1 — лише вертикаль.\n"
    "   - 3 vs 9, 5 vs 6: дивись нижні сегменти.\n"
    "7. Справжня кома відділяє останні 1-2 цифри. Читай ПОВНІСТЮ З КОМОЮ (напр. 048457.99).\n"
    "8. Якщо є блік/розмитість/сумнів у будь-якій цифрі — confidence=low.\n"
    'Поверни СТРОГО JSON без markdown: {"reading":"<повний показник з комою>","confidence":"high|low"}'
)


@app.post("/read")
def read(req: ReadRequest):
    if not ANTHROPIC_KEY:
        raise HTTPException(500, "ANTHROPIC_API_KEY не заданий")
    try:
        img = decode_img(req.image_base64)
        crop = crop_enhance(img, req.x1, req.y1, req.x2, req.y2, req.scale, req.clip_limit)
        ok, buf = cv2.imencode(".jpg", crop, [cv2.IMWRITE_JPEG_QUALITY, 90])
        crop_b64 = base64.b64encode(buf.tobytes()).decode()
    except Exception as e:
        raise HTTPException(500, f"crop error: {e}")

    anchor = ""
    if req.prev_reading:
        anchor = (f"КЛЮЧОВА ПІДКАЗКА: попередній показник був {req.prev_reading}. "
                  f"Новий МАЄ БУТИ близьким або трохи більшим (та сама довжина цілої частини). "
                  f"Якщо твій результат менший/коротший на розряд — ти згубив цифру, перечитай.\n")
    system = READ_SYS.replace("{ANCHOR}", anchor)
    try:
        txt = call_claude(system, crop_b64,
                          f"Зчитай повний показник з комою. Попередній був {req.prev_reading or '?'} — новий близький. Не загуби останню цифру.")
    except Exception as e:
        raise HTTPException(502, f"Claude read error: {e}")
    p = extract_json(txt)
    return {"reading": p.get("reading", ""), "confidence": p.get("confidence", "low")}


@app.get("/")
def health():
    return {"status": "ok", "service": "meter-ocr", "version": "3.3", "has_key": bool(ANTHROPIC_KEY)}


# ---------- /export ----------
class ExportRow(BaseModel):
    apartment: str = ""
    serial: str = ""
    prev_reading: str = ""
    reading_int: str = ""
    reading_full: str = ""
    diff: str = ""
    status: str = ""


class ExportRequest(BaseModel):
    rows: list[ExportRow]
    object_name: str = "Олімпійська 10Б"
    period: str = ""


def _apt_sort_key(a):
    try:
        return (0, int(str(a)))
    except Exception:
        return (1, str(a))


def _build_xlsx(rows, title, header_row, value_fn, sheet_name):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    import io as _io
    thin = Border(*[Side(style="thin")] * 4)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws["A1"] = title
    ws["A1"].font = Font(name="Arial", bold=True, size=11)
    ws.append(header_row)
    for c in range(1, len(header_row) + 1):
        cell = ws.cell(row=2, column=c)
        cell.font = Font(name="Arial", bold=True, size=10)
        cell.fill = PatternFill("solid", start_color="D9E1F2")
        cell.border = thin
        cell.alignment = Alignment(wrap_text=True, horizontal="center")
    srt = sorted(rows, key=lambda r: _apt_sort_key(r.apartment))
    for r in srt:
        ws.append(value_fn(r))
    for row in ws.iter_rows(min_row=3, max_row=2 + len(srt), min_col=1, max_col=len(header_row)):
        for cell in row:
            cell.border = thin
            cell.font = Font(name="Arial", size=10)
    for i, w in enumerate([12, 16, 18, 16, 14], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    buf = _io.BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode()


@app.post("/export")
def export(req: ExportRequest):
    def intval(v):
        try:
            return int(float(str(v)))
        except Exception:
            return ""
    # Бухгалтерія (цілі)
    buh = _build_xlsx(
        req.rows,
        f"Олимпийская э.счетчики {req.period}",
        ["номер кв", "Серійний номер", "Попередні покази", "Покази", "Різниця розрахована"],
        lambda r: [r.apartment, r.serial, intval(r.prev_reading), intval(r.reading_int), intval(r.diff)],
        "Квартиры и офисы",
    )
    # Обленерго (з десятими)
    obl = _build_xlsx(
        req.rows,
        f"{req.object_name} — покази для Обленерго, {req.period}",
        ["№ кв", "Серійний №", "Попередні покази", "Покази (з десятими)", "Різниця"],
        lambda r: [r.apartment, r.serial, intval(r.prev_reading), r.reading_full, intval(r.diff)],
        "Обленерго",
    )
    return {"buhgalteria_b64": buh, "oblenergo_b64": obl, "count": len(req.rows)}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=int(os.environ.get("PORT", 8000)))
